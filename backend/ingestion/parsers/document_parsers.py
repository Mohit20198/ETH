"""
Per-type document parsers.
Vaibhav owns this file. Each parser returns:
  - text: str (cleaned text for embedding)
  - chunks: list[str] (semantic chunks for embedding)
  - metadata: dict (doc-level metadata)
"""
import base64
import hashlib
import json
import os
import re
from pathlib import Path
from typing import Optional

import pdfplumber
import pytesseract
from PIL import Image
from docx import Document as DocxDocument
import openpyxl
from openai import OpenAI

# OpenCV is optional - P&ID parser degrades to OCR-only if not installed
try:
    import cv2
    import numpy as np
    OPENCV_AVAILABLE = True
except ImportError:
    OPENCV_AVAILABLE = False
    print("[Parsers] WARNING: opencv-python not installed. P&ID preprocessing disabled (OCR-only mode).")

from backend.config import settings

# Set Tesseract path if on Windows
if settings.TESSERACT_CMD:
    pytesseract.pytesseract.tesseract_cmd = settings.TESSERACT_CMD

_openai_client = OpenAI(api_key=settings.OPENAI_API_KEY)


# -----------------------------------------------------------------------------
# Document Fingerprinting (deduplication)
# -----------------------------------------------------------------------------

def fingerprint_file(file_path: str) -> str:
    """SHA-256 hash of file content for deduplication."""
    h = hashlib.sha256()
    with open(file_path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def load_fingerprint_cache(cache_path: str) -> dict:
    if os.path.exists(cache_path):
        with open(cache_path) as f:
            return json.load(f)
    return {}


def save_fingerprint_cache(cache: dict, cache_path: str):
    os.makedirs(os.path.dirname(cache_path) or ".", exist_ok=True)
    with open(cache_path, "w") as f:
        json.dump(cache, f, indent=2)


# -----------------------------------------------------------------------------
# Semantic Chunking (per doc type)
# -----------------------------------------------------------------------------

def semantic_chunk(text: str, doc_type: str, max_chars: int = None) -> list[str]:
    """
    Different chunking strategies per doc type.
    - SOPs/Procedures: chunk by numbered section
    - Work orders: chunk by work order block
    - Manuals: chunk by paragraph (800 chars)
    - Generic: fixed-size overlap chunks
    """
    if not text.strip():
        return []

    if doc_type in ("sop", "regulation", "manual"):
        sections = re.split(r'\n(?=\d+\.|[A-Z]\.|[IVX]+\.)', text)
        return [s.strip() for s in sections if len(s.strip()) > 50]

    elif doc_type == "work_order":
        orders = re.split(r'\n(?=WO[-\s]?\d+|Work Order:)', text, flags=re.IGNORECASE)
        return [o.strip() for o in orders if len(o.strip()) > 50]

    elif doc_type == "inspection":
        items = re.split(r'\n(?=\d+\s*[\.\)]\s+[A-Z])', text)
        return [i.strip() for i in items if len(i.strip()) > 30]

    else:
        size = max_chars or 800
        overlap = 100
        chunks = []
        start = 0
        while start < len(text):
            end = start + size
            chunks.append(text[start:end].strip())
            start = end - overlap
        return [c for c in chunks if len(c) > 50]


# -----------------------------------------------------------------------------
# PDF Parser
# -----------------------------------------------------------------------------

def parse_pdf(file_path: str, doc_type: str = "generic") -> dict:
    """Parse text PDFs using pdfplumber. Falls back to OCR for scanned pages."""
    text_pages = []
    metadata = {"page_count": 0}

    with pdfplumber.open(file_path) as pdf:
        metadata["page_count"] = len(pdf.pages)
        for i, page in enumerate(pdf.pages):
            page_text = page.extract_text() or ""
            if len(page_text.strip()) < 50:
                # Scanned page - use OCR
                img = page.to_image(resolution=200).original
                page_text = pytesseract.image_to_string(img)
            text_pages.append(page_text)

    full_text = "\n\n".join(text_pages)
    return {
        "text": full_text,
        "chunks": semantic_chunk(full_text, doc_type),
        "metadata": metadata,
    }


# -----------------------------------------------------------------------------
# Spreadsheet Parser
# -----------------------------------------------------------------------------

def parse_spreadsheet(file_path: str, doc_type: str = "spreadsheet") -> dict:
    """
    Parse Excel/CSV treating spatial layout as meaningful.
    Each sheet becomes a structured text block.
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    all_text = []
    metadata = {"sheets": []}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            row_vals = [str(cell) if cell is not None else "" for cell in row]
            if any(v.strip() for v in row_vals):
                rows.append(" | ".join(row_vals))

        if rows:
            sheet_text = f"## Sheet: {sheet_name}\n" + "\n".join(rows)
            all_text.append(sheet_text)
            metadata["sheets"].append(sheet_name)

    full_text = "\n\n".join(all_text)
    return {
        "text": full_text,
        "chunks": semantic_chunk(full_text, "spreadsheet"),
        "metadata": metadata,
    }


# -----------------------------------------------------------------------------
# P&ID / Engineering Drawing Parser
# -----------------------------------------------------------------------------

def parse_pid(file_path: str) -> dict:
    """
    P&ID parser:
    1. OpenCV - pre-process image (threshold, denoise) [if available]
    2. GPT-4V - identify equipment tags and rough connections

    IMPORTANT: On ANY vision API error, returns None — caller must treat None
    as extraction failure and must NOT write anything to vector store or graph.
    Falls back to OCR-only + GPT-4V on raw image if opencv not installed.
    """
    temp_path = None

    if OPENCV_AVAILABLE:
        img = cv2.imread(file_path)
        if img is None:
            return None  # Cannot load image — do not ingest
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        thresh = cv2.adaptiveThreshold(
            gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2
        )
        denoised = cv2.fastNlMeansDenoising(thresh, h=10)
        temp_path = file_path + "_preprocessed.png"
        cv2.imwrite(temp_path, denoised)
        ocr_text = pytesseract.image_to_string(denoised)
        vision_source_path = temp_path
    else:
        try:
            pil_img = Image.open(file_path)
            ocr_text = pytesseract.image_to_string(pil_img)
        except Exception as e:
            print(f"[P&ID Parser] OCR failed: {e}")
            ocr_text = ""
        vision_source_path = file_path

    # GPT-4V for symbol/tag detection
    # NOTE: max_tokens is NOT used — some model versions reject it (e.g. Luna)
    vision_result = None
    try:
        with open(vision_source_path, "rb") as f:
            img_b64 = base64.b64encode(f.read()).decode()

        response = _openai_client.chat.completions.create(
            model=settings.OPENAI_VISION_MODEL,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "text",
                            "text": (
                                "This is a P&ID (Piping and Instrumentation Diagram) from an industrial facility.\n"
                                "Extract ALL equipment tags (e.g. P-101, V-201, E-301, TI-101) and their rough connections.\n"
                                "Return JSON: {\"equipment_tags\": [...], \"connections\": [{\"from\": \"tag1\", \"to\": \"tag2\", \"line_type\": \"process/instrument/utility\"}]}\n"
                                "Focus on what is clearly readable. Do not guess."
                            ),
                        },
                        {
                            "type": "image_url",
                            "image_url": {"url": f"data:image/png;base64,{img_b64}", "detail": "high"},
                        },
                    ],
                }
            ],
        )
        raw_content = response.choices[0].message.content or ""

        # Guard: treat any response that looks like an error as a failure
        if not raw_content or any(
            marker in raw_content.lower()
            for marker in ["error code", "api error", "failed", "exception"]
        ):
            raise ValueError(f"Vision API returned suspicious content: {raw_content[:200]}")

        vision_result = raw_content

    except Exception as e:
        print(f"[P&ID Parser] Vision extraction FAILED for {file_path}: {e}")
        print(f"[P&ID Parser] Returning None — this document will NOT be ingested.")
    finally:
        if temp_path and os.path.exists(temp_path):
            os.remove(temp_path)

    # If vision failed AND we have no useful OCR text, refuse to store anything
    if vision_result is None:
        if not ocr_text.strip() or len(ocr_text.strip()) < 20:
            # Absolutely nothing useful to store — bail out cleanly
            return None

        # We have OCR text but no vision — store OCR-only with clear provenance
        pid_text = f"P&ID OCR (vision unavailable):\n{ocr_text.strip()}"
        extraction_status = "partial"
    else:
        pid_text = f"P&ID OCR:\n{ocr_text}\n\nVision Analysis:\n{vision_result}"
        extraction_status = "success"

    return {
        "text": pid_text,
        "chunks": [pid_text],
        "metadata": {
            "doc_type": "pid",
            "opencv_used": OPENCV_AVAILABLE,
            "extraction_status": extraction_status,
        },
    }


# -----------------------------------------------------------------------------
# Plain Text / Email / Report Parser
# -----------------------------------------------------------------------------

def parse_text(file_path: str, doc_type: str = "generic") -> dict:
    """Parse plain text, markdown, email files."""
    with open(file_path, "r", encoding="utf-8", errors="replace") as f:
        text = f.read()
    return {
        "text": text,
        "chunks": semantic_chunk(text, doc_type),
        "metadata": {"char_count": len(text)},
    }


# -----------------------------------------------------------------------------
# DOCX Parser
# -----------------------------------------------------------------------------

def parse_docx(file_path: str, doc_type: str = "generic") -> dict:
    doc = DocxDocument(file_path)
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    full_text = "\n\n".join(paragraphs)
    return {
        "text": full_text,
        "chunks": semantic_chunk(full_text, doc_type),
        "metadata": {"paragraph_count": len(paragraphs)},
    }


# -----------------------------------------------------------------------------
# Router - pick parser based on file extension + doc_type hint
# -----------------------------------------------------------------------------

def parse_document(file_path: str, doc_type: str = "generic") -> dict:
    ext = Path(file_path).suffix.lower()

    if doc_type == "pid" or ext in (".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp"):
        return parse_pid(file_path)
    elif ext == ".pdf":
        return parse_pdf(file_path, doc_type)
    elif ext in (".xlsx", ".xls", ".csv"):
        return parse_spreadsheet(file_path, doc_type)
    elif ext in (".docx", ".doc"):
        return parse_docx(file_path, doc_type)
    elif ext in (".txt", ".md", ".eml", ".msg"):
        return parse_text(file_path, doc_type)
    else:
        return parse_text(file_path, doc_type)
